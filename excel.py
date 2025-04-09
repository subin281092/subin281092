import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
import os

# Function to read column and row names from txt file
def read_labels_from_txt(file_path):
    try:
        with open(file_path, 'r') as f:
            lines = f.readlines()
        column_names, row_names = [], []

        for line in lines:
            if line.lower().startswith("columns:"):
                column_names = [name.strip() for name in line.split(":", 1)[1].split(",")]
            elif line.lower().startswith("rows:"):
                row_names = [name.strip() for name in line.split(":", 1)[1].split(",")]

        return row_names, column_names
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read file:\n{e}")
        return [], []

# Function to create the Excel file
def generate_excel(row_names, column_names, output_file='custom_labeled_sheet.xlsx'):
    try:
        wb = Workbook()
        ws = wb.active

        # Column headers
        for col_index, name in enumerate(column_names, start=2):
            ws.cell(row=1, column=col_index, value=name)

        # Row headers
        for row_index, name in enumerate(row_names, start=2):
            ws.cell(row=row_index, column=1, value=name)

        wb.save(output_file)
        messagebox.showinfo("Success", f"Excel file created:\n{output_file}")
    except Exception as e:
        messagebox.showerror("Error", f"Could not save Excel file:\n{e}")

# GUI handler
def on_generate_click():
    try:
        num_rows = int(row_entry.get())
        num_cols = int(col_entry.get())

        if num_rows <= 0 or num_cols <= 0:
            raise ValueError("Row and column counts must be positive integers.")

        file_path = filedialog.askopenfilename(
            title="Select Label TXT File",
            filetypes=[("Text Files", "*.txt")]
        )

        if not file_path:
            return

        row_names, column_names = read_labels_from_txt(file_path)

        if len(row_names) < num_rows or len(column_names) < num_cols:
            messagebox.showwarning("Warning", "Not enough labels in the file for the given dimensions.\nLabels will be truncated.")

        # Use only as many names as specified
        selected_rows = row_names[:num_rows]
        selected_cols = column_names[:num_cols]

        output_file = os.path.join(os.getcwd(), "generated_excel.xlsx")
        generate_excel(selected_rows, selected_cols, output_file)
    except ValueError as ve:
        messagebox.showerror("Input Error", str(ve))

# Build GUI
app = tk.Tk()
app.title("Excel Sheet Generator")
app.geometry("400x250")
app.resizable(False, False)

tk.Label(app, text="Enter number of rows:").pack(pady=5)
row_entry = tk.Entry(app)
row_entry.pack(pady=5)

tk.Label(app, text="Enter number of columns:").pack(pady=5)
col_entry = tk.Entry(app)
col_entry.pack(pady=5)

generate_btn = tk.Button(app, text="Upload File & Generate Excel", command=on_generate_click)
generate_btn.pack(pady=20)

app.mainloop()
